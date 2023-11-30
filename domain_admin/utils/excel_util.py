# -*- coding: utf-8 -*-
"""
@File    : excel_util.py
@Date    : 2023-11-16
"""
import re

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter


def read_excel(filename):
    """
    读取excel文件为python对象
    :param filename:
    :return: iterator
    """
    book = load_workbook(filename)
    worksheet = book.worksheets[0]

    titles = []
    row_num = 0

    for row in worksheet.rows:
        row_num += 1

        if row_num == 1:
            # 表头
            titles = [cell.value.strip() for cell in row]
        else:
            # 内容
            yield dict(zip(titles, [cell.value for cell in row]))

    book.close()


def write_excel(filename, rows):
    """
    将列表写入到文件
    :param filename:
    :param rows: list
    :return:
    """
    workbook = Workbook()
    worksheet = workbook.active


    # 表头
    if len(rows) > 0:
        for i, key in enumerate(rows[0].keys()):
            worksheet.cell(row=1, column=i + 1, value=key)

    # 内容
    for x, row in enumerate(rows):
        for y, value in enumerate(row.values()):
            worksheet.cell(row=x + 2, column=y + 1, value=value)

    # 调整列宽
    # 参考：https://blog.csdn.net/gongzairen/article/details/130819231
    width = 3  # 手动加宽的数值
    # 单元格列宽处理
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))

        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value + width

    workbook.save(filename)
    workbook.close()


if __name__ == '__main__':
    write_excel('./demo.xlsx', [
        {'name': 'Tom', 'age': 23},
        {'name': 'Jack', 'age': 24},
    ])
